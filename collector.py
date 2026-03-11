#!/usr/bin/env python3
"""
链接内容记录与分类工具 V1.1

功能：
1. 接收用户提交的链接或本地文件
2. 支持格式：网页、PDF、Excel
3. 使用 Playwright 爬取网页内容
4. 使用 PyPDF2/pdfplumber 提取 PDF 内容
5. 使用 openpyxl/pandas 提取 Excel 内容
6. 调用百炼 GLM-5 进行智能分类
7. 生成摘要和标签
8. 归档到 ideas-and-notes/inbox/
"""

import sys
import os
import json
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any

# 项目路径
PROJECT_ROOT = Path('/root/.openclaw/workspace/ideas-and-notes')
INBOX_DIR = PROJECT_ROOT / 'inbox'

# 分类标准
CATEGORIES = {
    'tech': '技术相关（编程、架构、工具）',
    'investment': '投资理财（股票、基金、经济）',
    'life': '生活日常（健康、旅行、美食）',
    'reading': '阅读笔记（书籍、文章、思考）',
    'tools': '工具资源（软件、服务、资源）'
}

# 智能分类 Prompt
CLASSIFICATION_PROMPT = """你是一个内容分类助手。请分析以下网页内容，并进行智能分类。

## 分类标准
{categories}

## 网页内容
标题: {title}
来源: {url}
内容摘要: {content}

## 输出要求
请以 JSON 格式输出：
{{
    "category": "分类（从分类标准中选择一个）",
    "importance": "必读/值得关注/参考",
    "tags": ["标签1", "标签2", "标签3"],
    "summary": "一句话摘要（不超过50字）",
    "key_points": ["要点1", "要点2", "要点3"]
}}

只输出 JSON，不要其他内容。"""


class LinkCollector:
    """链接内容收集器"""
    
    def __init__(self):
        # 使用百炼 GLM-5 配置
        self.config = self._load_config()
        self.base_url = self.config.get('baseUrl', 'https://coding.dashscope.aliyuncs.com/v1')
        self.api_key = self.config.get('apiKey', '')
        
    def _load_config(self) -> dict:
        """加载百炼配置"""
        config_path = Path('/root/.openclaw/openclaw.json')
        if config_path.exists():
            with open(config_path, 'r') as f:
                config = json.load(f)
            return config.get('models', {}).get('providers', {}).get('qwencode', {})
        return {}
    
    def extract_content_playwright(self, url: str, cookies: list = None) -> Dict[str, Any]:
        """
        使用 Playwright 提取网页内容
        
        Args:
            url: 网页链接
            cookies: Cookie 列表（可选）
            
        Returns:
            包含标题、内容的字典
        """
        try:
            from playwright.sync_api import sync_playwright
            
            with sync_playwright() as p:
                # 添加浏览器参数绕过检测
                browser = p.chromium.launch(
                    headless=True,
                    args=[
                        '--disable-blink-features=AutomationControlled',
                        '--disable-dev-shm-usage',
                        '--no-sandbox',
                        '--disable-setuid-sandbox',
                    ]
                )
                
                context = browser.new_context(
                    user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    viewport={'width': 1920, 'height': 1080}
                )
                
                # 添加 Cookie（如果提供）
                if cookies:
                    context.add_cookies(cookies)
                
                page = context.new_page()
                
                print(f"  正在爬取: {url}")
                page.goto(url, timeout=60000)
                
                # 等待页面加载
                page.wait_for_load_state('networkidle', timeout=30000)
                
                # 提取标题
                title = page.title() or '未知标题'
                
                # 提取正文内容
                # 尝试多种选择器
                content_selectors = [
                    'article',
                    '.article-content',
                    '.post-content',
                    '.content',
                    'main',
                    'body'
                ]
                
                content = ''
                for selector in content_selectors:
                    try:
                        element = page.query_selector(selector)
                        if element:
                            content = element.inner_text()
                            if len(content) > 200:
                                break
                    except:
                        continue
                
                if not content:
                    content = page.inner_text('body')
                
                browser.close()
                
                return {
                    'title': title,
                    'content': content[:10000],  # 限制长度
                    'url': url
                }
                
        except Exception as e:
            return {'error': f'Playwright 爬取失败: {e}'}
    
    def extract_content(self, url: str) -> Dict[str, Any]:
        """
        提取内容（自动识别链接或文件）
        
        Args:
            url: 网页链接或本地文件路径
            
        Returns:
            包含标题、内容的字典
        """
        # 检查是否为本地文件
        if url.startswith('/') or url.startswith('./') or url.startswith('~/'):
            return self.extract_local_file(url)
        
        # 检查是否为 file:// 协议
        if url.startswith('file://'):
            return self.extract_local_file(url[7:])
        
        # 网页链接
        cookies = self._load_cookies_for_url(url)
        result = self.extract_content_playwright(url, cookies)
        
        if 'error' not in result and result.get('content'):
            return result
        
        print("  Playwright 失败，尝试百炼联网搜索...")
        return self.extract_content_bailian(url)
    
    def extract_local_file(self, file_path: str) -> Dict[str, Any]:
        """
        提取本地文件内容
        
        Args:
            file_path: 本地文件路径
            
        Returns:
            包含标题、内容的字典
        """
        from pathlib import Path
        
        # 展开路径
        file_path = Path(file_path).expanduser().resolve()
        
        if not file_path.exists():
            return {'error': f'文件不存在: {file_path}'}
        
        # 获取文件扩展名
        suffix = file_path.suffix.lower()
        file_name = file_path.stem
        
        print(f"  处理本地文件: {file_path}")
        print(f"  文件类型: {suffix}")
        
        # 根据文件类型选择提取方法
        if suffix == '.pdf':
            return self.extract_pdf_content(file_path)
        elif suffix in ['.xlsx', '.xls']:
            return self.extract_excel_content(file_path)
        elif suffix in ['.md', '.txt']:
            return self.extract_text_content(file_path)
        else:
            return {'error': f'不支持的文件类型: {suffix}'}
    
    def extract_pdf_content(self, file_path: Path) -> Dict[str, Any]:
        """
        提取 PDF 文件内容
        
        Args:
            file_path: PDF 文件路径
            
        Returns:
            包含标题、内容的字典
        """
        try:
            # 尝试使用 pdfplumber（更准确）
            try:
                import pdfplumber
                
                text_content = []
                with pdfplumber.open(file_path) as pdf:
                    print(f"  PDF 页数: {len(pdf.pages)}")
                    for i, page in enumerate(pdf.pages[:20]):  # 最多提取前20页
                        text = page.extract_text()
                        if text:
                            text_content.append(f"--- 第{i+1}页 ---\n{text}")
                
                content = '\n\n'.join(text_content)
                
            except ImportError:
                # 回退到 PyPDF2
                import PyPDF2
                
                text_content = []
                with open(file_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    print(f"  PDF 页数: {len(reader.pages)}")
                    for i, page in enumerate(reader.pages[:20]):
                        text = page.extract_text()
                        if text:
                            text_content.append(f"--- 第{i+1}页 ---\n{text}")
                
                content = '\n\n'.join(text_content)
            
            if not content.strip():
                return {'error': 'PDF 内容提取为空'}
            
            # 使用文件名作为标题
            title = file_path.stem
            
            return {
                'title': title,
                'content': content[:20000],  # 限制长度
                'url': f'file://{file_path}',
                'file_type': 'pdf',
                'file_path': str(file_path)
            }
            
        except Exception as e:
            return {'error': f'PDF 提取失败: {e}'}
    
    def extract_excel_content(self, file_path: Path) -> Dict[str, Any]:
        """
        提取 Excel 文件内容
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            包含标题、内容的字典
        """
        try:
            # 尝试使用 pandas
            try:
                import pandas as pd
                
                # 读取所有工作表
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names
                
                print(f"  Excel 工作表数: {len(sheet_names)}")
                
                content_parts = []
                content_parts.append(f"## 文件: {file_path.name}")
                content_parts.append(f"## 工作表: {', '.join(sheet_names)}\n")
                
                for sheet_name in sheet_names[:5]:  # 最多处理前5个工作表
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    content_parts.append(f"\n### 工作表: {sheet_name}")
                    content_parts.append(f"行数: {len(df)}, 列数: {len(df.columns)}")
                    content_parts.append(f"列名: {', '.join(df.columns.astype(str))}")
                    
                    # 数据预览（前10行）
                    content_parts.append("\n数据预览:")
                    content_parts.append(df.head(10).to_string())
                    
                    # 数据统计
                    content_parts.append("\n数据统计:")
                    content_parts.append(df.describe(include='all').to_string())
                
                content = '\n'.join(content_parts)
                
            except ImportError:
                # 回退到 openpyxl
                from openpyxl import load_workbook
                
                wb = load_workbook(file_path, read_only=True, data_only=True)
                sheet_names = wb.sheetnames
                
                print(f"  Excel 工作表数: {len(sheet_names)}")
                
                content_parts = []
                content_parts.append(f"## 文件: {file_path.name}")
                content_parts.append(f"## 工作表: {', '.join(sheet_names)}\n")
                
                for sheet_name in sheet_names[:5]:
                    ws = wb[sheet_name]
                    
                    content_parts.append(f"\n### 工作表: {sheet_name}")
                    content_parts.append(f"行数: {ws.max_row}, 列数: {ws.max_column}")
                    
                    # 数据预览（前10行）
                    content_parts.append("\n数据预览:")
                    for row in ws.iter_rows(min_row=1, max_row=11, values_only=True):
                        content_parts.append(' | '.join(str(cell) if cell else '' for cell in row))
                
                content = '\n'.join(content_parts)
            
            # 使用文件名作为标题
            title = file_path.stem
            
            return {
                'title': title,
                'content': content[:20000],
                'url': f'file://{file_path}',
                'file_type': 'excel',
                'file_path': str(file_path),
                'sheet_count': len(sheet_names) if 'sheet_names' in dir() else 0
            }
            
        except Exception as e:
            return {'error': f'Excel 提取失败: {e}'}
    
    def extract_text_content(self, file_path: Path) -> Dict[str, Any]:
        """
        提取文本文件内容
        
        Args:
            file_path: 文本文件路径
            
        Returns:
            包含标题、内容的字典
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            title = file_path.stem
            
            return {
                'title': title,
                'content': content[:20000],
                'url': f'file://{file_path}',
                'file_type': 'text',
                'file_path': str(file_path)
            }
            
        except Exception as e:
            return {'error': f'文本提取失败: {e}'}
    
    def _load_cookies_for_url(self, url: str) -> list:
        """根据 URL 加载对应的 Cookie"""
        cookie_dir = Path('/root/.openclaw/workspace/link-collector/cookies')
        
        # 识别网站
        if 'xueqiu.com' in url:
            cookie_file = cookie_dir / 'xueqiu.json'
        elif 'dianping.com' in url:
            cookie_file = cookie_dir / 'dianping.json'
        elif 'xiaohongshu.com' in url:
            cookie_file = cookie_dir / 'xiaohongshu.json'
        else:
            return []
        
        # 加载 Cookie
        if cookie_file.exists():
            try:
                with open(cookie_file, 'r') as f:
                    return json.load(f)
            except:
                pass
        
        return []
    
    def extract_content_bailian(self, url: str) -> Dict[str, Any]:
        """
        使用百炼联网搜索提取内容（备选方案）
        """
        try:
            import httpx
            
            # 使用百炼 Chat API 配合 enable_search
            response = httpx.post(
                f"{self.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "qwen3-max-2026-01-23",  # 支持 enable_search 的模型
                    "messages": [{
                        "role": "user", 
                        "content": f"请提取这个网页的标题和正文内容：{url}\n\n请按以下格式输出：\n## 标题\n[网页标题]\n\n## 正文\n[网页正文内容]"
                    }],
                    "max_tokens": 4000,
                    "temperature": 0.7,
                    "enable_search": True
                },
                timeout=300
            )
            
            if response.status_code == 200:
                data = response.json()
                content = data['choices'][0]['message']['content']
                
                # 尝试从内容中提取标题
                lines = content.split('\n')
                title = lines[0] if lines else '未知标题'
                
                return {
                    'title': title.replace('#', '').strip(),
                    'content': content,
                    'url': url
                }
            else:
                return {'error': f'百炼请求失败: {response.status_code}'}
                
        except Exception as e:
            return {'error': f'百炼提取失败: {e}'}
    
    def classify_content(self, title: str, content: str, url: str) -> Dict[str, Any]:
        """
        智能分类内容（使用百炼 GLM-5）
        
        Args:
            title: 标题
            content: 内容
            url: 来源链接
            
        Returns:
            分类结果
        """
        import httpx
        
        categories_text = '\n'.join([f'- {k}: {v}' for k, v in CATEGORIES.items()])
        prompt = CLASSIFICATION_PROMPT.format(
            categories=categories_text,
            title=title,
            url=url,
            content=content[:3000]
        )
        
        try:
            response = httpx.post(
                f"{self.base_url}/chat/completions",
                headers={
                    "Authorization": f"Bearer {self.api_key}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "glm-5",
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 2000,
                    "temperature": 0.7
                },
                timeout=300
            )
            
            if response.status_code == 200:
                data = response.json()
                response_text = data['choices'][0]['message']['content']
                
                # 解析 JSON
                json_match = re.search(r'\{[\s\S]*\}', response_text)
                if json_match:
                    return json.loads(json_match.group())
            
            return {
                'category': 'reading',
                'importance': '值得关注',
                'tags': [],
                'summary': title,
                'key_points': []
            }
                
        except Exception as e:
            print(f"分类失败: {e}")
            return {
                'category': 'reading',
                'importance': '值得关注',
                'tags': [],
                'summary': title,
                'key_points': []
            }
    
    def save_to_inbox(self, url: str, title: str, content: str, 
                      classification: Dict[str, Any],
                      file_type: str = 'web') -> str:
        """
        保存到收件箱
        
        Args:
            url: 来源链接或文件路径
            title: 标题
            content: 内容
            classification: 分类结果
            file_type: 文件类型 (web/pdf/excel/text)
            
        Returns:
            保存的文件路径
        """
        # 创建日期目录
        today = datetime.now().strftime('%Y-%m-%d')
        inbox_date_dir = INBOX_DIR / today
        inbox_date_dir.mkdir(parents=True, exist_ok=True)
        
        # 生成文件名（使用时间戳保证唯一性）
        timestamp = datetime.now().strftime('%H%M%S')
        safe_title = re.sub(r'[^\w\s-]', '', title)[:30]
        safe_title = re.sub(r'[-\s]+', '-', safe_title)
        
        # 保存原始内容
        raw_filename = f"{timestamp}_{safe_title}_raw.md"
        raw_filepath = inbox_date_dir / raw_filename
        raw_filepath.write_text(content, encoding='utf-8')
        
        # 生成归档文件名
        archive_filename = f"{timestamp}_{safe_title}.md"
        filepath = inbox_date_dir / archive_filename
        
        # 生成 Markdown 内容
        # 文件类型图标
        file_type_icons = {
            'web': '🌐',
            'pdf': '📄',
            'excel': '📊',
            'text': '📝'
        }
        file_type_icon = file_type_icons.get(file_type, '📎')
        
        md_content = f"""# {title}

## 元数据

| 属性 | 值 |
|------|------|
| **来源** | {file_type_icon} [{url}]({url}) |
| **类型** | {file_type.upper()} |
| **分类** | {classification.get('category', 'reading')} |
| **重要性** | {classification.get('importance', '值得关注')} |
| **标签** | {', '.join(classification.get('tags', []))} |
| **采集时间** | {datetime.now().strftime('%Y-%m-%d %H:%M')} |
| **原始内容** | [{raw_filename}](./{raw_filename}) |

## 摘要

{classification.get('summary', '暂无摘要')}

## 要点

"""
        for point in classification.get('key_points', []):
            md_content += f"- {point}\n"
        
        if not classification.get('key_points'):
            md_content += "*要点提取中...*\n"
        
        md_content += f"""
## 内容预览

{content[:500]}...

---
*由 link-collector V1.1 自动采集*
"""
        
        # 写入文件
        filepath.write_text(md_content, encoding='utf-8')
        
        return filepath, raw_filepath
    
    def process_link(self, url: str) -> Dict[str, Any]:
        """
        处理单个链接或文件
        
        Args:
            url: 网页链接或本地文件路径
            
        Returns:
            处理结果
        """
        print(f"正在处理: {url}")
        
        # 1. 提取内容
        extracted = self.extract_content(url)
        if 'error' in extracted:
            return extracted
        
        title = extracted.get('title', '未知标题')
        content = extracted.get('content', '')
        file_type = extracted.get('file_type', 'web')
        
        print(f"  标题: {title}")
        print(f"  内容长度: {len(content)} 字符")
        if file_type != 'web':
            print(f"  文件类型: {file_type}")
        
        # 2. 智能分类
        print("  正在分类...")
        classification = self.classify_content(title, content, url)
        
        print(f"  分类: {classification.get('category')}")
        print(f"  重要性: {classification.get('importance')}")
        print(f"  标签: {classification.get('tags')}")
        
        # 3. 保存到收件箱
        filepath, raw_filepath = self.save_to_inbox(url, title, content, classification, file_type)
        
        print(f"  已保存: {filepath}")
        print(f"  原始内容: {raw_filepath}")
        
        return {
            'success': True,
            'title': title,
            'category': classification.get('category'),
            'importance': classification.get('importance'),
            'tags': classification.get('tags'),
            'summary': classification.get('summary'),
            'filepath': str(filepath),
            'raw_filepath': str(raw_filepath),
            'file_type': file_type
        }


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("=" * 60)
        print("链接内容记录与分类工具 V1.1")
        print("=" * 60)
        print("\n用法: python collector.py <URL或文件路径>")
        print("\n支持格式:")
        print("  🌐 网页链接 - https://example.com/article")
        print("  📄 PDF文件  - /path/to/file.pdf")
        print("  📊 Excel文件 - /path/to/file.xlsx")
        print("  📝 文本文件  - /path/to/file.md")
        print("\n示例:")
        print("  python collector.py https://xueqiu.com/123456")
        print("  python collector.py ~/Downloads/report.pdf")
        print("  python collector.py ./data.xlsx")
        return
    
    url = sys.argv[1]
    
    collector = LinkCollector()
    result = collector.process_link(url)
    
    print("\n" + "="*60)
    print("处理结果:")
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == '__main__':
    main()